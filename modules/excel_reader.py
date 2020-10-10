# Author: Dada Liu (劉永琦)
import openpyxl
import os

def get_address_list(path, start_row, col_name):
    wb = openpyxl.load_workbook(path)
    ws = wb[wb.sheetnames[0]]
    address_list = list()
    for i in range(start_row, ws.max_row+1):
        tmp_addr = ws[f"{col_name}{i}"].value
        address = get_pure_address(tmp_addr)
        address_list.append(address)
    return address_list

def get_pure_address(address):
    # 過濾郵遞區號
    idx = 0
    for ch in address:
        if ch.isdigit(): 
            idx += 1
        else: 
            break
    address = address[idx:]
        
    # 過濾地址非必要字串
    address = address[:address.find("號")+1]
    if "/" in address:
        address = address.replace("/", "")
    return address
    
if __name__ == "__main__":
    pass
    # 測試: 
    '''
    path = os.listdir("../raw_excel_data")[0] # 第一份excel文件
    start_row = 5  # 第一筆資料所在的列數(起始row)
    col_name = "K" # "通訊地址"所在column
    
    address_list = get_address_list(path, start_row, col_name)
    print(address_list)
    '''